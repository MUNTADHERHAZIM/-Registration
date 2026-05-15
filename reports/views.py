from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.db.models import Q, Count, Avg
from students.models import Student, Department, Semester, AcademicYear
from core.models import SystemLog
from django.contrib.auth.models import User
from django.utils import timezone
from datetime import timedelta
import json
import io
from arabic_reshaper import reshape
from bidi.algorithm import get_display

def fix_arabic(text):
    if not text:
        return ""
    return get_display(reshape(str(text)))

@login_required
def admin_performance_dashboard(request):
    if not request.user.is_staff:
        return redirect('dashboard')
        
    now = timezone.now()
    today = now.date()
    
    # 1. User Productivity (Based on SystemLog)
    # We look for "إضافة طالب جديد" in details to be accurate
    user_actions = SystemLog.objects.filter(
        action='create', 
        details__contains='إضافة طالب جديد'
    ).values('user__username', 'user__first_name', 'user__last_name').annotate(
        count=Count('id')
    ).order_by('-count')
    
    # 2. Time-based Trends (Daily, Weekly, Monthly)
    daily_count = Student.objects.filter(created_at__date=today).count()
    weekly_count = Student.objects.filter(created_at__gte=now - timedelta(days=7)).count()
    monthly_count = Student.objects.filter(created_at__gte=now - timedelta(days=30)).count()
    
    # 3. Department Performance (New students in last 30 days)
    dept_performance = Department.objects.annotate(
        new_students=Count('students', filter=Q(students__created_at__gte=now - timedelta(days=30)))
    ).values('name', 'new_students').order_by('-new_students')
    
    # 4. Governorate Distribution (Geographical)
    gov_stats = Student.objects.values('governorate').annotate(
        count=Count('id')
    ).order_by('-count')
    
    # 5. Last 7 Days Activity (For Chart)
    days_list = []
    daily_stats = []
    for i in range(6, -1, -1):
        d = today - timedelta(days=i)
        days_list.append(d.strftime('%Y-%m-%d'))
        daily_stats.append(Student.objects.filter(created_at__date=d).count())

    chart_data = {
        'days': days_list,
        'daily_counts': daily_stats,
        'user_names': [u['user__first_name'] or u['user__username'] for u in user_actions],
        'user_counts': [u['count'] for u in user_actions],
        'gov_labels': [g['governorate'] or 'غير محدد' for g in gov_stats],
        'gov_counts': [g['count'] for g in gov_stats],
    }

    context = {
        'title': 'لوحة مراقبة الأداء الإداري',
        'user_actions': user_actions,
        'daily_count': daily_count,
        'weekly_count': weekly_count,
        'monthly_count': monthly_count,
        'dept_performance': dept_performance,
        'gov_stats': gov_stats,
        'chart_data_json': json.dumps(chart_data),
    }
    return render(request, 'reports/admin_performance.html', context)


def generate_pdf_header(canvas_obj, doc, title, subtitle=''):
    from reportlab.lib.units import cm
    from reportlab.lib import colors
    canvas_obj.saveState()
    canvas_obj.setFillColor(colors.HexColor('#1a237e'))
    canvas_obj.rect(0, doc.pagesize[1] - 3*cm, doc.pagesize[0], 3*cm, fill=1)
    canvas_obj.setFillColor(colors.white)
    canvas_obj.setFont('Helvetica-Bold', 18)
    canvas_obj.drawCentredString(doc.pagesize[0]/2, doc.pagesize[1] - 1.5*cm, title)
    if subtitle:
        canvas_obj.setFont('Helvetica', 11)
        canvas_obj.drawCentredString(doc.pagesize[0]/2, doc.pagesize[1] - 2.2*cm, subtitle)
    canvas_obj.restoreState()


@login_required
def reports_home(request):
    return render(request, 'reports/reports_home.html', {'title': 'التقارير والإحصاءات'})


@login_required
def students_report(request):
    dept_id = request.GET.get('department', '')
    level = request.GET.get('level', '')
    status = request.GET.get('status', 'active')
    gender = request.GET.get('gender', '')

    students = Student.objects.select_related('department', 'entry_year')
    if dept_id:
        students = students.filter(department_id=dept_id)
    if level:
        students = students.filter(level=level)
    if status:
        students = students.filter(status=status)
    if gender:
        students = students.filter(gender=gender)

    selected_fields = request.GET.getlist('fields')
    if not selected_fields:
        selected_fields = ['student_id', 'full_name', 'department', 'level', 'status', 'gpa']

    export = request.GET.get('export', '')
    if export == 'pdf':
        return export_students_pdf(students, request, selected_fields)
    if export == 'excel':
        return export_students_excel(students, selected_fields)
    if export == 'print':
        return render(request, 'reports/students_report_print.html', {
            'students': students,
            'selected_fields': selected_fields,
            'title': 'تقرير الطلاب التفصيلي',
        })

    stats = {
        'total': students.count(),
        'male': students.filter(gender='male').count(),
        'female': students.filter(gender='female').count(),
        'avg_gpa': students.aggregate(avg=Avg('gpa'))['avg'] or 0,
    }

    all_fields_list = [
        ('student_id', 'الرقم الجامعي'),
        ('full_name', 'الاسم الكامل'),
        ('department', 'القسم'),
        ('level', 'المرحلة'),
        ('gender', 'الجنس'),
        ('status', 'الحالة'),
        ('gpa', 'المعدل'),
        ('phone', 'الهاتف'),
        ('guardian_phone', 'ولي الأمر'),
        ('admission_channel', 'قناة القبول'),
        ('entry_year', 'سنة القبول'),
        ('study_type', 'نوع الدراسة'),
    ]

    return render(request, 'reports/students_report.html', {
        'students': students[:100],
        'stats': stats,
        'departments': Department.objects.all(),
        'dept_id': dept_id,
        'level': level,
        'status': status,
        'gender': gender,
        'selected_fields': selected_fields,
        'all_fields_list': all_fields_list,
        'title': 'تقرير الطلاب',
    })


def export_students_pdf(students, request, selected_fields):
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_RIGHT

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4),
                            rightMargin=1.5*cm, leftMargin=1.5*cm,
                            topMargin=4*cm, bottomMargin=2*cm)
    styles = getSampleStyleSheet()
    elements = []

    # Title
    title_style = ParagraphStyle('title', parent=styles['Title'], fontSize=16, spaceAfter=20, alignment=TA_CENTER)
    elements.append(Paragraph('تقرير الطلاب', title_style))
    elements.append(Paragraph(f'عدد الطلاب: {students.count()}', styles['Normal']))
    elements.append(Spacer(1, 0.5*cm))

    # Table headers and data mapping
    all_fields = {
        'student_id': 'الرقم الجامعي',
        'full_name': 'الاسم الكامل',
        'department': 'القسم',
        'level': 'المرحلة',
        'gender': 'الجنس',
        'status': 'الحالة',
        'gpa': 'المعدل',
        'phone': 'رقم الهاتف',
        'guardian_phone': 'رقم ولي الأمر',
        'admission_channel': 'قناة القبول',
        'entry_year': 'سنة القبول',
    }
    
    headers = ['#'] + [all_fields[f] for f in selected_fields if f in all_fields]
    data = [headers]
    
    for i, s in enumerate(students, 1):
        row = [str(i)]
        for f in selected_fields:
            if f == 'student_id': row.append(s.student_id)
            elif f == 'full_name': row.append(s.full_name)
            elif f == 'department': row.append(s.department.name)
            elif f == 'level': row.append(s.get_level_display())
            elif f == 'gender': row.append(s.get_gender_display())
            elif f == 'status': row.append(s.get_status_display())
            elif f == 'gpa': row.append(str(s.gpa))
            elif f == 'phone': row.append(s.phone or '')
            elif f == 'guardian_phone': row.append(s.guardian_phone or '')
            elif f == 'admission_channel': row.append(s.get_admission_channel_display() if hasattr(s, 'get_admission_channel_display') else s.admission_channel or '')
            elif f == 'entry_year': row.append(s.entry_year.year)
        data.append(row)

    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a237e')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTSIZE', (0, 0), (-1, 0), 11),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f5f5f5')]),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#dddddd')),
        ('PADDING', (0, 0), (-1, -1), 6),
    ]))
    elements.append(table)
    doc.build(elements)
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="students_report.pdf"'
    return response


def export_students_excel(students, selected_fields):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'الطلاب'
    ws.sheet_view.rightToLeft = True

    header_fill = PatternFill(start_color='1a237e', end_color='1a237e', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=12)
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    all_fields = {
        'student_id': 'الرقم الجامعي',
        'full_name': 'الاسم الكامل',
        'department': 'القسم',
        'level': 'المرحلة',
        'gender': 'الجنس',
        'status': 'الحالة',
        'gpa': 'المعدل التراكمي',
        'phone': 'رقم الهاتف',
        'guardian_phone': 'رقم ولي الأمر',
        'admission_channel': 'قناة القبول',
        'entry_year': 'سنة القبول',
        'study_type': 'نوع الدراسة',
    }
    
    headers = ['#'] + [all_fields[f] for f in selected_fields if f in all_fields]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

    for row_idx, s in enumerate(students, 2):
        row_data = [row_idx-1]
        for f in selected_fields:
            if f == 'student_id': row_data.append(s.student_id)
            elif f == 'full_name': row_data.append(s.full_name)
            elif f == 'department': row_data.append(s.department.name)
            elif f == 'level': row_data.append(s.get_level_display())
            elif f == 'gender': row_data.append(s.get_gender_display())
            elif f == 'status': row_data.append(s.get_status_display())
            elif f == 'gpa': row_data.append(float(s.gpa))
            elif f == 'phone': row_data.append(s.phone or '')
            elif f == 'guardian_phone': row_data.append(s.guardian_phone or '')
            elif f == 'admission_channel': row_data.append(s.get_admission_channel_display() if hasattr(s, 'get_admission_channel_display') else s.admission_channel or '')
            elif f == 'entry_year': row_data.append(s.entry_year.year)
            elif f == 'study_type': row_data.append(s.get_study_type_display())
        
        for col_idx, v in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=v)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
            if row_idx % 2 == 0:
                cell.fill = PatternFill(start_color='f5f5f5', end_color='f5f5f5', fill_type='solid')

    for col in ws.columns:
        max_len = max(len(str(c.value or '')) for c in col)
        ws.column_dimensions[col[0].column_letter].width = max(12, max_len + 4)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="students_report.xlsx"'
    return response


@login_required
def statistics_report(request):
    import json
    
    # 1. Department Stats
    dept_stats = Department.objects.annotate(
        total=Count('students'),
        active=Count('students', filter=Q(students__status='active')),
        graduated=Count('students', filter=Q(students__status='graduated')),
        avg_gpa=Avg('students__gpa'),
        avg_admission=Avg('students__admission_avg'),
    )
    
    # 2. Basic Distributions
    gender_stats = list(Student.objects.values('gender').annotate(count=Count('id')))
    level_stats = list(Student.objects.values('level').annotate(count=Count('id')).order_by('level'))
    branch_stats = list(Student.objects.values('branch').annotate(count=Count('id')).order_by('-count'))[:10]
    channel_stats = list(Student.objects.values('admission_channel').annotate(count=Count('id')).order_by('-count'))[:8]
    governorate_stats = list(Student.objects.values('governorate').annotate(count=Count('id')).order_by('-count'))[:10]
    origin_stats = list(Student.objects.values('origin_status').annotate(count=Count('id')))
    marital_stats = list(Student.objects.values('marital_status').annotate(count=Count('id')))
    round_stats = list(Student.objects.values('admission_round').annotate(count=Count('id')).order_by('admission_round'))
    religion_stats = list(Student.objects.values('religion').annotate(count=Count('id')).order_by('-count'))

    # 3. Top Students
    top_students = Student.objects.order_by('-admission_avg')[:10]
    
    # 4. Numeric Averages
    overall_stats = Student.objects.aggregate(
        avg_admission=Avg('admission_avg'),
        avg_no_additions=Avg('avg_no_additions'),
        total=Count('id')
    )

    # 5. Prepare Chart Data
    chart_data = {
        'dept': {
            'labels': [d.name for d in dept_stats],
            'counts': [d.total for d in dept_stats]
        },
        'gender': {
            'labels': [dict(Student.GENDER_CHOICES).get(g['gender'], g['gender']) for g in gender_stats],
            'counts': [g['count'] for g in gender_stats]
        },
        'branch': {
            'labels': [b['branch'] or 'غير محدد' for b in branch_stats],
            'counts': [b['count'] for b in branch_stats]
        },
        'channel': {
            'labels': [c['admission_channel'] or 'غير محدد' for c in channel_stats],
            'counts': [c['count'] for c in channel_stats]
        },
        'gov': {
            'labels': [g['governorate'] or 'غير محدد' for g in governorate_stats],
            'counts': [g['count'] for g in governorate_stats]
        },
        'origin': {
            'labels': [o['origin_status'] or 'غير محدد' for o in origin_stats],
            'counts': [o['count'] for o in origin_stats]
        },
        'round': {
            'labels': [f'الدور {r["admission_round"]}' if r["admission_round"] else 'غير محدد' for r in round_stats],
            'counts': [r['count'] for r in round_stats]
        }
    }

    context = {
        'dept_stats': dept_stats,
        'overall_stats': overall_stats,
        'top_students': top_students,
        'religion_stats': religion_stats,
        'marital_stats': marital_stats,
        'chart_data_json': json.dumps(chart_data),
        'title': 'الإحصاءات الشاملة للطلاب',
    }
    return render(request, 'reports/statistics_report.html', context)

